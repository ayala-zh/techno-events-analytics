import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

def export_to_excel(dataframes_dict, filename):
    # Ensure exports folder exists
    export_dir = "exports"
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, filename)

    # Write DataFrames to Excel
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        total_rows = 0
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            total_rows += len(df)

    # Load workbook with openpyxl for formatting
    wb = load_workbook(filepath)

    for sheet_name in dataframes_dict.keys():
        ws = wb[sheet_name]

        # Freeze header row + first column
        ws.freeze_panes = "B2"

        # Apply filters to all columns
        ws.auto_filter.ref = ws.dimensions

        # Apply conditional formatting to numeric columns
        for col in ws.iter_cols(min_row=2, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            # Check if column is numeric (based on first non-empty cell)
            if all((isinstance(cell.value, (int, float)) or cell.value is None) for cell in col):
                col_letter = col[0].column_letter
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

    wb.save(filepath)

    # Print console message
    print(f"Created file {filename}, {len(dataframes_dict)} sheets, {total_rows} rows")
